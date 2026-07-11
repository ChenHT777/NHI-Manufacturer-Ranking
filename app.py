import os
import re
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# --- 網頁基本設定 (必須放在腳本最上方) ---
st.set_page_config(page_title="健保資料庫數據分析工具", page_icon="💊", layout="wide")

# --- 1. 固定讀取同目錄下的資料庫檔案 ---
DATA_FILE = "data.csv"

@st.cache_data # 加入快取機制，大幅提升網頁流暢度
def load_fixed_database():
    if os.path.exists(DATA_FILE):
        try:
            df = pd.read_csv(DATA_FILE)
            df.columns = df.columns.str.strip()
            for col in ['成分', '劑型', '劑量', '廠商', '年度']:
                if col in df.columns:
                    df[col] = df[col].astype(str).str.strip()
            return df
        except Exception as e:
            st.error(f"讀取本地資料庫失敗: {e}")
            return pd.DataFrame()
    else:
        return pd.DataFrame()

df_global = load_fixed_database()

if not df_global.empty:
    ALL_COMPONENTS = sorted(df_global['成分'].dropna().unique())
    INIT_MESSAGE = f"✅ 成功載入資料庫！共 {len(df_global)} 筆數據，包含 {len(ALL_COMPONENTS)} 種成分。"
else:
    ALL_COMPONENTS = []
    INIT_MESSAGE = "❌ 未找到 data.csv 檔案，請將資料庫檔案上傳至同目錄。"

# --- 2. 輔助函式 ---
def parse_dosage_to_numeric(dose_str):
    if not isinstance(dose_str, str):
        return 0.0
    match = re.search(r'([0-9\.]+)', dose_str)
    if match:
        val = float(match.group(1))
        if 'g' in dose_str.lower() and 'mg' not in dose_str.lower() and 'mcg' not in dose_str.lower():
            return val * 1000
        return val
    return 0.0

# --- 3. 產出標準化 Excel 與 HTML 預覽報表 ---
def generate_reports(selected_components, selected_forms, selected_doses):
    if df_global.empty:
        return None, "❌ 系統未成功載入 data.csv 資料庫。", ""
    
    if not selected_components or not selected_forms:
        return None, "❌ 請確認成分與劑型皆已勾選！", ""

    condition = (df_global['成分'].isin(selected_components)) & (df_global['劑型'].isin(selected_forms))
    if selected_doses:
        condition &= (df_global['劑量'].isin(selected_doses))

    df_filtered = df_global[condition].copy()
    
    if df_filtered.empty:
        return None, "❌ 找不到符合該條件的資料！", ""

    df_filtered['劑量'] = df_filtered['劑量'].astype(str).str.strip()
    df_filtered['劑量'] = df_filtered['劑量'].apply(
        lambda x: '' if str(x).lower() in ['nan', 'none', '<na>', 'null', ''] else x
    )

    qty_col = '數量(顆)' if '數量(顆)' in df_filtered.columns else [col for col in df_filtered.columns if '數量' in col][0]
    df_filtered[qty_col] = df_filtered[qty_col].astype(str).str.replace(',', '').str.strip()
    df_filtered[qty_col] = pd.to_numeric(df_filtered[qty_col], errors='coerce').fillna(0)

    pivot_df = df_filtered.groupby(['成分', '劑型', '劑量', '廠商', '年度'])[qty_col].sum().unstack(fill_value=0)
    for year_col in ['2022年', '2023年', '2024年']:
        if year_col not in pivot_df.columns:
            pivot_df[year_col] = 0
            
    pivot_df = pivot_df.reindex(columns=['2022年', '2023年', '2024年']).reset_index()
    pivot_df['dose_numeric'] = pivot_df['劑量'].apply(parse_dosage_to_numeric)
    pivot_df = pivot_df.sort_values(
        by=['成分', '劑型', 'dose_numeric', '2024年'], 
        ascending=[True, True, True, False]
    ).drop(columns=['dose_numeric'])

    # --- Excel 處理邏輯 ---
    wb = Workbook()
    ws = wb.active
    ws.title = "廠商排名報表"
    ws.views.sheetView[0].showGridLines = True
    ws.views.sheetView[0].zoomScale = 85

    font_family = "微軟正黑體"
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    subtotal_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    total_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    
    header_font = Font(name=font_family, size=12, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=12)
    bold_font = Font(name=font_family, size=12, bold=True)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    
    thin_side = Side(border_style="thin", color="D9D9D9")
    thick_bottom_side = Side(border_style="medium", color="000000")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    thick_bottom_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom_side)

    # --- HTML 處理邏輯 ---
    html_content = f"""
    <style>
        .report-preview-wrapper {{ padding: 15px 0; }}
        .report-container {{ 
            background-color: #FFFFFF !important; color: #333333 !important; padding: 20px !important; 
            width: 100%; box-sizing: border-box; -webkit-text-size-adjust: 100%;
            border-radius: 12px; box-shadow: 0 8px 24px rgba(0,0,0,0.25); border: 1px solid #E0E0E0;
        }}
        .scroll-hint {{
            display: none; text-align: right; font-size: 13px; color: #888888 !important; 
            margin-bottom: 8px; font-family: '微軟正黑體', sans-serif; font-weight: bold; animation: pulse 2s infinite;
        }}
        @keyframes pulse {{ 0% {{ opacity: 0.5; }} 50% {{ opacity: 1; }} 100% {{ opacity: 0.5; }} }}
        @media (max-width: 768px) {{ .scroll-hint {{ display: block; }} }}
        .table-responsive {{ width: 100%; overflow-x: auto; -webkit-overflow-scrolling: touch; padding-bottom: 10px; }}
        .table-responsive::-webkit-scrollbar {{ height: 10px; }}
        .table-responsive::-webkit-scrollbar-track {{ background: #F1F1F1; border-radius: 5px; }}
        .table-responsive::-webkit-scrollbar-thumb {{ background: #C0C0C0; border-radius: 5px; }}
        .report-table {{ width: max-content; min-width: 100%; border-collapse: collapse; font-family: '微軟正黑體', sans-serif; font-size: 15px !important; background-color: #FFFFFF !important; }}
        .report-table th, .report-table td {{ border: 1px solid #D9D9D9 !important; padding: 8px 12px; white-space: nowrap !important; vertical-align: middle !important; color: #333333 !important; }}
        .report-table th {{ background-color: #1F497D !important; color: #FFFFFF !important; text-align: center !important; font-weight: bold; }}
        .report-title {{ color: #000000 !important; text-align: center; font-family: '微軟正黑體'; font-weight: bold; margin-top: 5px; margin-bottom: 20px; }}
        .thick-bottom {{ border-bottom: 2px solid black !important; }}
        .report-footer {{ display: flex; flex-wrap: wrap; justify-content: space-between; margin-top: 25px; padding-top: 15px; border-top: 1px solid #D9D9D9; font-size: 13px !important; font-family: '微軟正黑體', sans-serif; font-weight: bold; color: #333333 !important; }}
    </style>
    <div class="report-preview-wrapper"><div id="report-capture-area" class="report-container">
    """

    comp_names = "、".join(df_filtered['成分'].unique())
    form_mapping = {"注射劑": "Inj.", "一般錠劑膠囊劑": "Tab./Cap.", "膜衣錠": "F.C. Tab.", "膠囊劑": "Cap.", "錠劑": "Tab."}
    unique_forms = df_filtered['劑型'].unique()
    form_abbr = f" {form_mapping.get(unique_forms[0], unique_forms[0])}" if len(unique_forms) == 1 else ""
    
    valid_doses = [d for d in df_filtered['劑量'].unique() if d != '']
    doses_for_header = sorted(valid_doses, key=parse_dosage_to_numeric)
    doses_str = "、".join(doses_for_header)
    dose_display = f" {doses_str}" if doses_str else ""
    
    header_title_text = f"{comp_names}{form_abbr}{dose_display}廠商申報量排名"

    html_content += f'<h2 class="report-title">{header_title_text}</h2>'
    html_content += '<div class="scroll-hint">👉 左右滑動表格可查看完整數據</div>'
    html_content += '<div class="table-responsive"><table class="report-table">'

    headers = ["成分", "劑型", "劑量", "廠商", "2022年<br>數量", "2023年<br>數量", "2024年<br>數量", "2024年<br>占比(%)"]
    ws_headers = [h.replace("<br>", "\n") for h in headers]
    ws.append(ws_headers)
    ws.row_dimensions[1].height = 45 
    
    html_content += "<tr>"
    for col_idx, h in enumerate(ws_headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = cell_border
        html_content += f"<th style='color: #FFFFFF !important; background-color: #1F497D !important;'>{headers[col_idx-1]}</th>"
    html_content += "</tr>"

    unique_groups = pivot_df[['成分', '劑型', '劑量']].drop_duplicates()
    unique_groups_list = list(unique_groups.iterrows())
    
    grand_total_2022 = grand_total_2023 = grand_total_2024 = 0
    current_row = 2
    last_comp = last_form = None

    for idx, (_, group_keys) in enumerate(unique_groups_list):
        comp = group_keys['成分']
        form = group_keys['劑型']
        dose = group_keys['劑量']

        is_last_of_form = False
        if idx == len(unique_groups_list) - 1:
            is_last_of_form = True
        else:
            next_comp = unique_groups_list[idx+1][1]['成分']
            next_form = unique_groups_list[idx+1][1]['劑型']
            if comp != next_comp or form != next_form: is_last_of_form = True

        dose_group = pivot_df[(pivot_df['成分'] == comp) & (pivot_df['劑型'] == form) & (pivot_df['劑量'] == dose)]
        
        dose_2024_sum = dose_group['2024年'].sum()
        dose_subtotal_2022 = dose_group['2022年'].sum()
        dose_subtotal_2023 = dose_group['2023年'].sum()
        dose_subtotal_2024 = dose_2024_sum
        
        print_comp = (comp != last_comp)
        print_form = print_comp or (form != last_form)
        is_first_row_in_dose = True
        
        for _, row in dose_group.iterrows():
            qty_2022, qty_2023, qty_2024 = float(row['2022年']), float(row['2023年']), float(row['2024年'])
            ratio = (qty_2024 / dose_2024_sum) if dose_2024_sum > 0 else 0.0
            
            c_val = row['成分'] if print_comp and is_first_row_in_dose else ""
            f_val = row['劑型'] if print_form and is_first_row_in_dose else ""
            d_val = row['劑量'] if is_first_row_in_dose else ""
             
            ws.append([c_val, f_val, d_val, row['廠商'], qty_2022, qty_2023, qty_2024, ratio])
            ws.row_dimensions[current_row].height = 35
            
            html_content += f"<tr><td style='text-align:center;'>{c_val}</td><td style='text-align:center;'>{f_val}</td><td style='text-align:center;'>{d_val}</td><td style='text-align:left;'>{row['廠商']}</td><td style='text-align:right;'>{qty_2022:,.0f}</td><td style='text-align:right;'>{qty_2023:,.0f}</td><td style='text-align:right;'>{qty_2024:,.0f}</td><td style='text-align:right;'>{ratio:.1%}</td></tr>"
            
            for col_idx in range(1, 9):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = data_font; cell.border = cell_border
                if col_idx in [1, 2, 3]: cell.alignment = center_align
                elif col_idx == 4: cell.alignment = left_align
                elif col_idx in [5, 6, 7]: cell.alignment = right_align; cell.number_format = '#,##0'
                elif col_idx == 8: cell.alignment = right_align; cell.number_format = '0.0%'
            
            current_row += 1
            is_first_row_in_dose = False
        
        last_comp = comp
        last_form = form
            
        dose_label = f"{dose} 合計" if dose != "" else "合計"
        ws.append(["", "", dose_label, "", dose_subtotal_2022, dose_subtotal_2023, dose_subtotal_2024, 1.0])
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=4)
        ws.row_dimensions[current_row].height = 35
        
        thick_class = "thick-bottom" if is_last_of_form else ""
        html_content += f"""
        <tr style="background-color: #DCE6F1; font-weight: bold;">
            <td class="{thick_class}"></td><td class="{thick_class}"></td>
            <td colspan="2" class="{thick_class}" style="text-align:center;">{dose_label}</td>
            <td class="{thick_class}" style="text-align:right;">{dose_subtotal_2022:,.0f}</td>
            <td class="{thick_class}" style="text-align:right;">{dose_subtotal_2023:,.0f}</td>
            <td class="{thick_class}" style="text-align:right;">{dose_subtotal_2024:,.0f}</td>
            <td class="{thick_class}" style="text-align:right;">100.0%</td>
        </tr>
        """
        
        for col_idx in range(1, 9):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = bold_font; cell.fill = subtotal_fill; cell.border = thick_bottom_border if is_last_of_form else cell_border
            if col_idx == 3: cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            elif col_idx in [5, 6, 7]: cell.alignment = right_align; cell.number_format = '#,##0'
            elif col_idx == 8: cell.alignment = right_align; cell.number_format = '0.0%'
                
        grand_total_2022 += dose_subtotal_2022
        grand_total_2023 += dose_subtotal_2023
        grand_total_2024 += dose_subtotal_2024
        current_row += 1

    ws.append(["總計", "", "", "", grand_total_2022, grand_total_2023, grand_total_2024, 1.0])
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    ws.row_dimensions[current_row].height = 35
    
    html_content += f"""
    <tr style="background-color: #B8CCE4; font-weight: bold;">
        <td colspan="4" style="text-align:center;">總計</td>
        <td style="text-align:right;">{grand_total_2022:,.0f}</td>
        <td style="text-align:right;">{grand_total_2023:,.0f}</td>
        <td style="text-align:right;">{grand_total_2024:,.0f}</td>
        <td style="text-align:right;">100.0%</td>
    </tr></table></div>
    <div class="report-footer">
        <span>中央健康保險署 政府資料開放平台 2024年資料</span><span>https://data.gov.tw/dataset/22131</span>
    </div></div></div>
    """

    for col_idx in range(1, 9):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = bold_font; cell.fill = total_fill; cell.border = cell_border
        if col_idx == 1: cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        elif col_idx in [5, 6, 7]: cell.alignment = right_align; cell.number_format = '#,##0'
        elif col_idx == 8: cell.alignment = right_align; cell.number_format = '0.0%'

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter == 'A': ws.column_dimensions[col_letter].width = 35; continue
        elif col_letter == 'B': ws.column_dimensions[col_letter].width = 20; continue
        elif col_letter in ['E', 'F', 'G']: ws.column_dimensions[col_letter].width = 18; continue
        elif col_letter == 'H': ws.column_dimensions[col_letter].width = 11.5; continue
            
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            for line in val_str.split('\n'):
                line_len = sum(2 if '\u4e00' <= char <= '\u9fff' else 1 for char in line)
                if line_len > max_len: max_len = line_len
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0; ws.page_setup.paperSize = 9 
    ws.page_margins.top = 2.7 / 2.54; ws.page_margins.bottom = 2.5 / 2.54
    ws.page_margins.left = 1.5 / 2.54; ws.page_margins.right = 1.5 / 2.54
    ws.page_margins.header = 1.5 / 2.54; ws.page_margins.footer = 1.0 / 2.54

    ws.oddHeader.center.text = f'&"微軟正黑體,Bold"&16{header_title_text}'
    ws.oddFooter.left.text = '&"微軟正黑體,Regular"&12中央健康保險署 政府資料開放平台 2024年資料'
    ws.oddFooter.right.text = '&"微軟正黑體,Regular"&12https://data.gov.tw/dataset/22131'
    ws.page_setup.scaleWithDoc = True; ws.page_setup.alignWithMargins = True

    # 引入時間套件來產生流水號
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 擷取藥品名稱（成分），並過濾掉不合法的檔名字元
    comp_names_clean = "_".join(df_filtered['成分'].unique())
    comp_names_clean = re.sub(r'[\\/*?:"<>| ]', "_", comp_names_clean)
    
    # 組合出統一的主檔名：藥品名稱_時間流水號 (例如: Ibuprofen_20260711_153025)
    base_filename = f"{comp_names_clean}_{timestamp}"
    excel_filename = f"{base_filename}.xlsx"
    
    # 儲存 Excel
    wb.save(excel_filename)
    
    success_msg = "🎉 成功！已順利產出「Excel 報表」與下方「預覽畫面」。請點擊按鈕下載檔案。"
    # 👇 這裡多回傳一個 base_filename 給外層使用
    return excel_filename, success_msg, html_content, base_filename


# --- 4. 介面層 (Streamlit UI) ---

def check_password():
    """回傳 True 代表密碼正確，否則回傳 False"""

    # 檢查 session_state 中是否已經有驗證成功的紀錄
    if st.session_state.get("password_correct", False):
        return True

    # 若尚未驗證，則顯示密碼輸入介面
    st.title("🔒 系統登入")
    
    # 使用 st.form 將輸入框與按鈕綁定在一起
    # 這樣只要在 text_input 裡面按 Enter，就會自動觸發 form_submit_button
    with st.form("login_form"):
        password_input = st.text_input("請輸入系統存取密碼", type="password")
        submit_btn = st.form_submit_button("登入")
        
        # 當使用者點擊登入按鈕或按 Enter 鍵時，執行以下檢查
        if submit_btn:
            if password_input == st.secrets["app_password"]:
                st.session_state["password_correct"] = True
                st.rerun()  # 密碼正確，立刻重新載入畫面以顯示主程式
            else:
                st.error("❌ 密碼錯誤，請重新輸入。")
                
    return False

# -----------------------------------------
# 🚨 密碼攔截點：如果密碼不對，程式就在這裡停止運行，不往下執行
# -----------------------------------------
if not check_password():
    st.stop()
    
# =========================================
# 以下為原本的系統主畫面 (完全不需要修改，只要放在 check_password() 下方即可)
# =========================================

st.title("💊 健保資料庫數據分析工具")
st.markdown(f"**系統狀態：** {INIT_MESSAGE}")

# 建立左右兩欄
col1, col2 = st.columns(2)

with col1:
    st.markdown("### 🔍 輸入條件")
    
    # 第一步：搜尋並選擇成分
    selected_components = st.multiselect(
        "📋 第一步：搜尋並選擇成分 (支援打字即時搜尋)", 
        options=ALL_COMPONENTS, 
        placeholder="請輸入成分關鍵字（打越多字越精準，例如打 Levo...）"
    )
    
    # 第二步：動態產生劑型與全選邏輯
    form_options = []
    selected_forms = []
    if selected_components and not df_global.empty:
        df_filtered = df_global[df_global['成分'].isin(selected_components)]
        form_options = sorted(df_filtered['劑型'].dropna().unique())
        
        if form_options:
            select_all_forms = st.toggle("☑️ 第二步：全選所有劑型", value=True)
            
            if select_all_forms:
                selected_forms = form_options
            else:
                selected_forms = st.multiselect(
                    "💊 請手動選擇特定劑型", 
                    options=form_options, 
                    placeholder="請點擊此處展開選項..."
                )
    
    # 第三步：動態產生劑量與全選邏輯
    dose_options = []
    selected_doses = []
    if selected_components and selected_forms and not df_global.empty:
        df_filtered = df_global[
            (df_global['成分'].isin(selected_components)) & 
            (df_global['劑型'].isin(selected_forms))
        ]
        dose_options = sorted(df_filtered['劑量'].dropna().unique(), key=parse_dosage_to_numeric)
        
        if dose_options:
            select_all_doses = st.toggle("☑️ 第三步：全選所有劑量", value=True)
            
            if select_all_doses:
                selected_doses = dose_options
            else:
                selected_doses = st.multiselect(
                    "🧪 請手動選擇特定劑量", 
                    options=dose_options, 
                    placeholder="請點擊此處展開選項..."
                )
                
    # 👇 就是剛剛不小心漏掉的這行！把它補回來，右邊的程式碼就不會報錯了
    submit_btn = st.button("🚀 第四步：產生報表與預覽", type="primary", use_container_width=True)

with col2:
    st.markdown("### 📥 報表下載區")
    
    if submit_btn:
        with st.spinner("報表產生中，請稍候..."):
            # 👇 這裡改為接收 4 個回傳值
            file_name, msg, html_res, base_name = generate_reports(selected_components, selected_forms, selected_doses)
            
            if file_name:
                st.success(msg)
                # 👇 將主檔名紀錄到系統變數中，供最下方的圖片下載按鈕讀取
                st.session_state['current_base_filename'] = base_name
                
                with open(file_name, "rb") as f:
                    st.download_button(
                        label="📄 一鍵下載 Excel 報表",
                        data=f,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                st.session_state['html_preview'] = html_res
            else:
                st.error(msg)
    else:
        st.info("👈 請先於左側設定條件並點擊「產生報表」")

st.markdown("---")
st.markdown("### 網頁即時預覽")

# 顯示 HTML 預覽與一鍵圖片下載 (透過內嵌沙盒繞過限制)
# 顯示 HTML 預覽與一鍵圖片下載
# 👇 這裡多加一個判斷，確保檔名變數已經產生
if 'html_preview' in st.session_state and 'current_base_filename' in st.session_state:
    wrapped_html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
        <style>
            body {{ font-family: '微軟正黑體', sans-serif; margin: 0; padding: 10px; }}
            .download-btn-container {{ text-align: right; margin-bottom: 15px; }}
            .dl-btn {{
                background-color: #FF4B4B; color: white; border: none; padding: 10px 20px;
                border-radius: 8px; font-size: 15px; font-weight: bold; cursor: pointer;
                box-shadow: 0 4px 6px rgba(0,0,0,0.1); transition: 0.2s;
            }}
            .dl-btn:hover {{ background-color: #FF3333; transform: translateY(-2px); }}
        </style>
    </head>
    <body>
        <div class="download-btn-container">
            <button class="dl-btn" onclick="downloadImage()">📸 一鍵下載為高畫質圖片 (PNG)</button>
        </div>
        
        <div id="capture-target">
            {st.session_state['html_preview']}
        </div>

        <script>
            function downloadImage() {{
                var btn = document.querySelector('.dl-btn');
                var element = document.getElementById('capture-target');
                var wrapper = document.querySelector('.table-responsive');
                var scrollHint = element.querySelector('.scroll-hint');

                var originalWidth = element.style.width;
                var originalOverflow = wrapper ? wrapper.style.overflowX : '';
                var originalHintDisplay = scrollHint ? scrollHint.style.display : '';
                var originalBtnHtml = btn ? btn.innerHTML : '';

                var targetWidth = element.scrollWidth;
                element.style.width = targetWidth + 'px';
                if(wrapper) wrapper.style.overflowX = 'visible';
                if(scrollHint) scrollHint.style.display = 'none';

                // 避免手機瀏覽器（尤其 iOS）canvas 尺寸過大造成渲染失敗或當機
                var MAX_DIMENSION = 4000;
                var scale = (targetWidth * 2 > MAX_DIMENSION) ? Math.max(1, MAX_DIMENSION / targetWidth) : 2;

                if (btn) {{ btn.disabled = true; btn.innerHTML = '⏳ 圖片產生中，請稍候...'; }}

                function restoreUI() {{
                    element.style.width = originalWidth;
                    if(wrapper) wrapper.style.overflowX = originalOverflow;
                    if(scrollHint) scrollHint.style.display = originalHintDisplay;
                    if (btn) {{ btn.disabled = false; btn.innerHTML = originalBtnHtml; }}
                }}

                html2canvas(element, {{
                    scale: scale,
                    backgroundColor: '#FFFFFF',
                    useCORS: true,
                    width: targetWidth,
                    windowWidth: targetWidth,
                    scrollX: 0,
                    scrollY: 0
                }}).then(function(canvas) {{
                    restoreUI();

                    var fileName = '{st.session_state["current_base_filename"]}.png';

                    // 使用 Blob 而非 dataURL，避免手機瀏覽器因巨大字串而卡頓或當機
                    canvas.toBlob(function(blob) {{
                        if (!blob) {{
                            alert('圖片產生失敗，請稍後再試。');
                            return;
                        }}

                        var isIOS = /iPad|iPhone|iPod/.test(navigator.userAgent) || (navigator.platform === 'MacIntel' && navigator.maxTouchPoints > 1);
                        var isAndroid = /Android/.test(navigator.userAgent);
                        var isMobile = isIOS || isAndroid;

                        // 👇 關鍵改動：手機（含 iPhone 上的 Chrome/Safari）優先呼叫原生「分享」選單，
                        // 可直接選擇「儲存圖片」，比 <a download> 或長按更穩定可靠
                        if (isMobile && navigator.canShare) {{
                            try {{
                                var file = new File([blob], fileName, {{ type: 'image/png' }});
                                if (navigator.canShare({{ files: [file] }})) {{
                                    navigator.share({{ files: [file], title: fileName }}).catch(function(err) {{
                                        console.log('分享已取消或失敗:', err);
                                    }});
                                    return;
                                }}
                            }} catch (e) {{
                                console.warn('此瀏覽器不支援 Web Share，改用備援方案:', e);
                            }}
                        }}

                        var blobUrl = URL.createObjectURL(blob);

                        if (isIOS) {{
                            // 【iOS 備援方案】若無法呼叫分享選單，將圖片直接渲染在網頁上，引導長按儲存
                            var existingFallback = document.getElementById('ios-fallback-container');
                            if (existingFallback) existingFallback.remove();

                            var container = document.createElement('div');
                            container.id = 'ios-fallback-container';
                            container.style.marginTop = '25px';
                            container.style.padding = '15px';
                            container.style.border = '2px dashed #FF4B4B';
                            container.style.borderRadius = '10px';
                            container.style.textAlign = 'center';
                            container.style.backgroundColor = '#FFF6F6';

                            var msg = document.createElement('p');
                            msg.innerHTML = '📱 <b>iPhone / iPad 用戶請注意：</b><br>請👉 <b>長按下方圖片</b> 👈，選擇「儲存圖片」或「分享」即可！';
                            msg.style.color = '#333';
                            msg.style.fontSize = '16px';
                            msg.style.lineHeight = '1.6';
                            msg.style.fontFamily = '微軟正黑體, sans-serif';

                            var img = document.createElement('img');
                            img.src = blobUrl;
                            img.style.maxWidth = '100%';
                            img.style.border = '1px solid #CCC';
                            img.style.marginTop = '15px';
                            img.style.boxShadow = '0 4px 12px rgba(0,0,0,0.15)';

                            container.appendChild(msg);
                            container.appendChild(img);
                            document.body.appendChild(container);

                            // 自動將畫面往下捲動，讓使用者立刻看到生成的圖片
                            container.scrollIntoView({{behavior: 'smooth', block: 'end'}});

                        }} else {{
                            // 【Android 與電腦版】維持正常的自動下載功能
                            var link = document.createElement('a');
                            link.download = fileName;
                            link.href = blobUrl;
                            document.body.appendChild(link);
                            link.click();
                            document.body.removeChild(link);
                            setTimeout(function() {{ URL.revokeObjectURL(blobUrl); }}, 10000);
                        }}
                    }}, 'image/png');

                }}).catch(function(error) {{
                    restoreUI();
                    console.error('截圖失敗:', error);
                    alert('截圖產生失敗，請稍後再試。');
                }});
            }}
        </script>
    </body>
    </html>
    """
    
    components.html(wrapped_html, height=800, scrolling=True)